try:
    import pyodbc
    import pandas as pd
    import openpyxl
    import os
    import glob
    from openpyxl import load_workbook
    from openpyxl.styles import colors, PatternFill, Border, Side                      #To import PatternFill module
    from openpyxl.styles.differential import DifferentialStyle                         #To import DifferentialStyle module
    from openpyxl.formatting.rule import Rule                                          #To import Rule module
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
    from datetime import date, time, datetime, timedelta           #To import datetime modules date, time, and datetime
    import tkinter as tk
    from tkinter import messagebox
    from tkinter import *
except:
    import tkinter as tk
    from tkinter import messagebox
    from tkinter import *
    window = tk.Tk()
    window.withdraw()
    if messagebox.askyesno('Error', "Modules Missing!\nDOWNLOAD Required Modules?") == True:
        print("Installing Modules...")
        import os
        cmd = 'py -m pip install pyodbc openpyxl pandas'            #cmd is the variable that will store our shell command    
        upgrade = 'py -m pip install --upgrade pip'
        fp = os.popen(cmd)   #To launch the above command. The argument is a string that contains a shell command.
        print(fp.read())     #To read the output of the object
        finished = fp.close()   #command to close fp pipeline
  
        window.deiconify()
        window.destroy()
        window.quit()
        
    import pyodbc
    import pandas as pd
    import openpyxl
    import os
    import glob
    from openpyxl import load_workbook
    from openpyxl.styles import colors, PatternFill, Border, Side                      #To import PatternFill module
    from openpyxl.styles.differential import DifferentialStyle  #To import DifferentialStyle module
    from openpyxl.formatting.rule import Rule                   #To import Rule module
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
    from datetime import date, time, datetime, timedelta           #To import datetime modules date, time, and datetime
    import tkinter as tk
    from tkinter import messagebox
    from tkinter import *
    
    
    
#1. DIR
#2. DATETIME AND NEW FILE NAME
#3. MIGRATION
#4. CONDITIONAL FORMATING
#5. BORDERS ON ERRORED CELL


"""--------------------------------------------------------------------DIR----------------------------------------------------------------------------------------"""
cwd = os.getcwd()              #To get current working directory
all_files = os.listdir(cwd)    #To list all files in directory

if not any(fname.endswith('.accdb') for fname in os.listdir('.')):      #To check if current working directory contains MS Access file
    window = tk.Tk()
    window.withdraw()
    messagebox.showerror('Error', "No Microsoft Access file found in current working directory")
    quit()                  #To exit program

sorted_by_modification_time_ascending = sorted(all_files, key=lambda t: os.stat(t).st_mtime) #To sort files by recently modified

file = ""                       #Blank variable to store MS Access file

for i in sorted_by_modification_time_ascending:
    if ".accdb" in i:
        file = i                #To assign the most recently modified .accdb to variable called file
        
print (file)                    #To print the file name thatr we'll be working with

file_dir = os.path.join(cwd, file) #To get full path of file
#print (file_dir)

check_drivers = [i for i in pyodbc.drivers() if i.startswith('Microsoft Access Driver')]

if "*.accdb" in check_drivers[0]:
    conn_str = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'r'DBQ=' + file_dir + ';'    #New Driver (64-bit) – Microsoft Access Driver (*.mdb, *.accdb): works with 64-bit Python
elif "*.accdb" not in check_drivers[0]:
    conn_str = r'DRIVER={Microsoft Access Driver (*.mdb)};'r'DBQ=' + file_dir + ';'             #Old Driver (32-bit) – Microsoft Access Driver (*.mdb): works with 32-bit Python
else:
    window = tk.Tk()
    window.withdraw()
    messagebox.showerror('Error', "AccessDatabaseEngine_X64.exe Not Installed!")
    if messagebox.askyesno('Error', "DOWNLOAD AccessDatabaseEngine_X64.exe from https://www.microsoft.com?") == True:
        cmd = 'start msedge https://www.microsoft.com/en-US/download/details.aspx?id=13255' #To open website that hosts MS Access Databasee
        fp = os.popen(cmd)              #To launch the above command. The argument is a string that contains a shell command.
        print(fp.read())                #To read the output of the object
        finished = fp.close()           #command to close fp pipeline    
        window.deiconify()
        window.destroy()
        window.quit()
        quit()
    else:
        quit()


#"""----------------------------------------------------------DATETIME AND NEW FILE NAME----------------------------------------------------------------------------"""
today = date.today()                #Creating instance of today's date. It is in datetime.date format

new_file_name = 'Checked_' + str(today) + '.xlsx' #Creating new file output name

#"""------------------------------------------------------------------MIGRATION-------------------------------------------------------------------------------------"""
conn = pyodbc.connect(conn_str)
cursor = conn.cursor()              #“cursor” is an object used to execute SQL statements.

                    
sql = 'SELECT * FROM "U=U Table" ORDER BY "Enr date" DESC'  #To query the MS Access database

try:
    cursor.execute(sql)             #To run SQL statement    
except:
    window = tk.Tk()
    window.withdraw()
    messagebox.showerror('Error', '"U=U Table" not found in MS Access file')

df = pd.read_sql(sql, conn)         #Using Pandas to Query

try:
    df.to_excel (new_file_name)         #Using Pandas to export file in xlsx format
except:
    window = tk.Tk()
    #window.eval('tk::PlaceWindow %s center' % window.winfo_toplevel())
    window.withdraw()
    messagebox.showwarning('Error', f"Permission Denied. Please close {new_file_name}")
    window.deiconify()
    window.destroy()
    window.quit()
    quit ()
    
workbook = load_workbook(filename = new_file_name)  #Load workbook
sheet = workbook.active             #Load sheet


sheet.delete_cols(1)                #To clean the data/delete unnecessary columns

workbook.save(new_file_name)        #To save cleaned data


#"""------------------------------------------------------------------CONDITIONAL FORMATING-------------------------------------------------------------------------"""

workbook = load_workbook(filename = new_file_name) #To load workbook
sheet = workbook.active             #To load sheet
sheet.title = "DQA"


#Defining background colors
orange_background = PatternFill(bgColor="FFC300")     #For orange backgroung
red_background = PatternFill(bgColor="FF0202")        #For red backgroung
green_background = PatternFill(bgColor="1BBF01")      #For green backgroung
white_background = PatternFill(bgColor="FFFFFF")      #For white backgroung

#White background for headings
diff_style = DifferentialStyle(fill=white_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['$A1 = "Enr date"'] 
sheet.conditional_formatting.add(sheet.dimensions, rule)

#Flag Orange
diff_style = DifferentialStyle(fill=orange_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['OR($T1 = "Not Harmonized",$T1 = "", $U1 = "Not Harmonized", $U1 = "", AND($B1 = "Site Normal Enr",$AA1 = "No"), AND($B1 = "Community Normal Enr",$AA1 = "No"), AND($B1 = "Site Normal Enr",$AB1 = "NO"), AND($B1 = "Community Normal Enr",$AB1 = "NO"), $AD1 = "")'] 
sheet.conditional_formatting.add(sheet.dimensions, rule)

#Flag RED for Dates
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['OR($A1 > TODAY(), $A1 < DATE(2021, 1, 1))']
sheet.conditional_formatting.add(sheet.dimensions, rule)

#Flag RED for entries that have dropdown option but did entry is not amoung dropdowns (Also flags blank cells where dropdown is applicable)
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['OR(AND($B1 <> "Site Care Card Enr", $B1 <> "Site Normal Enr", $B1 <> "Communty Care Card Enr", $B1 <> "Community Normal Enr"), AND($D1 <> "Chisanga UHC", $D1 <> "Kasama General", $D1 <> "Kaizya HP", $D1 <> "Kasakalawe HP", $D1 <> "Mpulungu HAHC", $D1 <> "Nsumbu RHC", $D1 <> "Tulemane UHC"), AND($E1 <> "MCH", $E1 <> "Labour Ward", $E1 <> "VCT", $E1 <> "PITC", $E1 <> "DCT", $E1 <> "Fast Track", $E1 <> "Traige", $E1 <> "Youth Conner", $E1 <> "OPD", $E1 <> "VMMC", $E1 <> "Indexing", $E1 <> "ART", $E1 <> "T.B", $E1 <> "Community", $E1 <> "Cervical Cancer", $E1 <> "Mens Clinic", $E1 <> "Pediatric Ward", $E1 <> "PNS"), AND($M1 <> "Male", $M1 <> "Female"), AND($P1 <> "Active", $P1 <> "Inactive"), AND($Q1 <> "Trans Out", $Q1 <> "Trans In", $Q1 <> "Deceased", $Q1 <> "LTFU", $Q1 <> "Deactivated", $Q1 <> "Local"),  AND($AB1 <> "YES", $AB1 <> "NO", $AB1 <> "NO Mobile #"), AND($AA1 <> "Yes", $AA1 <> "No"), AND($N1 <> "English", $N1 <> "Nyanja", $N1 <> "Bemba", $N1 <> "Lungu", $N1 <> "Mambwe"), AND($S1 <> "New", $S1 <> "Old"), AND($BG1 <> "Fishermen/women", $BG1 <> "Farmers", $BG1 <> "Traders", $BG1 <> "Others", $BG1 <> ""), AND($U1 <> "Same in SC", $U1 <> "Same in PRs", $U1 <> "Same in Both", $U1 <> "Different or No Address in PRs/Added", $U1 <> "Different or No Address in Sc/Added", $U1 <> "Different or No Address in Both/Added", $U1 <> "Same in SC but Different or No Address in PRs", $U1 <> "Same in PRs but Different or No Address in SC", $U1 <> "Not Harmonized"), AND($AX1 <> "Yes", $AX1 <> "No"), AND($T1 <> "Same in SC", $T1 <> "Same in PRs", $T1 <> "Same in Both", $T1 <> "Different or No Mobile in PRs/Added", $T1 <> "Different or No Mobile in Sc/Added", $T1 <> "Different or No Mobile in Both/Added", $T1 <> "Same in SC but Different or No Mobile in PRs", $T1 <> "Same in PRs but Different or No Mobile in SC", $T1 <> "Care Card", $T1 <> "Not Harmonized"), AND($AW1 <> "Not Eligible (TX_NEW)", $AW1 <> "Results Found in SC and Updated in CTS", $AW1 <> "Results Found in Physical Registers", $AW1 <> "No VL Result found in SC or PRs", $AW1 <> "VL Updated after follow up", $AW1 <> "VL Results Pending Collection and Updates", $AW1 <> ""))']
sheet.conditional_formatting.add(sheet.dimensions, rule)


#Flag RED for impossible dates
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['OR($A1 > TODAY(), $A1 < 01/01/2020, AND($C1 <> "", $C1 < 1900-01-01), AND($C1 <> "", $C1 > TODAY()), $L1 < 01/01/1900, $L1 >= TODAY(), $AD1 > (TODAY() + 217), AND($AC1 <> "", $AC1 < (TODAY() - 217)), AND($AE1 <> "", $AE1 > (TODAY() + 217)), AND($AQ1 <> "", $AQ1 > (TODAY() + 730)), AND($BH1 <> "", $BH1 > TODAY()), AND($BH1 <> "", $AQ1 <> "", $BH1 = $AQ1), AND($BH1 <> "", $AQ1 <> "", $BH1 > $AQ1), AND($BD1 <> "", $BD1 > TODAY()))']
sheet.conditional_formatting.add(sheet.dimensions, rule)


#Flag RED for blank cells/Missing information (Excludes options with dropdowns because they are accomodated in previous constraint)
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['OR($A1 = "", $F1 = "", $G1 = "", $I1 = "", $J1 = "", $L1 = "", $R1 = "", AND($AW1 = "", AND($T1 <> "Not Harmonized", $U1 <> "Not Harmonized")))']
sheet.conditional_formatting.add(sheet.dimensions, rule)


#Flag RED for Status with inconsistent status comment
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['OR(AND($P1 = "Active", AND($Q1 <> "Local", Q1 <> "Trans In")), AND($P1 = "Inactive", OR($Q1 <> "Trans Out", Q1 = "Deceased", Q1 = "LTFU", Q1 = "Deactivated")))']#, AND($P1 = "Inactive", AND($Q1 <> "Local", $Q1 <> "Trans In")))']  
sheet.conditional_formatting.add(sheet.dimensions, rule)

#Flag RED for inconsistent Care Card Enrollment entries
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['OR(AND($B1 = "Site Care Card Enr", $AB1 <> "NO Mobile #"), AND($B1 = "Communty Care Card Enr", $AB1 <> "NO Mobile #"), AND($B1 = "Site Care Card Enr", $AA1 <> "No"), AND($B1 = "Communty Care Card Enr", $AA1 <> "No"), sAND($B1 = "Site Care Card Enr", $X1 <> ""), AND($B1 = "Site Care Card Enr", $Y1 <> ""), AND($B1 = "Site Care Card Enr", $Z1 <> ""), AND($B1 = "Communty Care Card Enr", $X1 <> ""), AND($B1 = "Communty Care Card Enr", $Y1 <> ""), AND($B1 = "Communty Care Card Enr", $Z1 <> ""), AND($B1 = "Site Care Card Enr", AND($T1 <> "Care Card", $T1 <> "Not Harmonized")), AND($B1 = "Communty Care Card Enr", AND($T1 <> "Care Card", $T1 <> "Not Harmonized")))']
sheet.conditional_formatting.add(sheet.dimensions, rule)


#Flag RED for inconsistent Normal Enrollment entries
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['OR(AND(AND($B1 = "Site Normal Enr", $X1 = ""), AND($B1 = "Site Normal Enr", $Y1 = ""), AND($B1 = "Site Normal Enr", $Z1 = "")), AND(AND($B1 = "Community Normal Enr", $X1 = ""), AND($B1 = "Community Normal Enr", $Y1 = ""), AND($B1 = "Community Normal Enr", $Z1 = "")))']#, AND(AND($B1 = "Community Normal Enr", $X1 = ""), AND($B1 = "Community Normal Enr", $Y1 = ""), AND($B1 = "Community Normal Enr", $Z1 = "")))']
sheet.conditional_formatting.add(sheet.dimensions, rule)

"""
#Flag RED for inconsistent Client type entries
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['OR(AND($S1 = "Old", $AW1 = "Not Eligible (TX_NEW)"))']
sheet.conditional_formatting.add(sheet.dimensions, rule)
"""

#Flag RED for inconsistent VL entries
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['OR(AND($AW1 = "Results Found in SC and Updated in CTS", $AR1 = "", $AS1 = ""), AND($AW1 = "Results Found in Physical Registers", $AR1 = "", $AS1 = ""), AND($AW1 = "Results Found in SC and Updated in CTS", $BH1 =""), AND($AW1 = "Results Found in Physical Registers", $BH1 =""))']
sheet.conditional_formatting.add(sheet.dimensions, rule)


#Flag RED for inconsistent Mobile numbers
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['OR(AND(LEFT($X1,2) <> "", LEFT($X1,2) <> "97", LEFT($X1,2) <> "77"), AND(LEFT($Y1,2) <> "", LEFT($Y1,2) <> "95", LEFT($Y1,2) <> "75"), AND(LEFT($Z1,2) <> "", LEFT($Z1,2) <> "96", LEFT($Z1,2) <> "76"))']
sheet.conditional_formatting.add(sheet.dimensions, rule)


#Flag RED for inconsistent entries associated with residential address
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['OR(AND($U1 = "Not Harmonized", $AX1 = "Yes"), AND($U1 = "", $AX1 = "Yes"))']
sheet.conditional_formatting.add(sheet.dimensions, rule)


#Flag GREEN for entries with no violation
diff_style = DifferentialStyle(fill=green_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ['AND($T1 <> "Not Harmonized",$T1 <> "", $U1 <> "Not Harmonized", $U1 <> "", OR(AND($B1 = "Site Normal Enr",$AA1 <> "No"), AND($B1 = "Community Normal Enr",$AA1 <> "No")), OR(AND($B1 = "Site Normal Enr",$AB1 <> "NO"), AND($B1 = "Community Normal Enr",$AB1 <> "NO")), $AD1 <> "", $A1 < TODAY(), $A1 > DATE(2021, 1, 1), OR($B1 = "Site Care Card Enr", $B1 = "Site Normal Enr", $B1 = "Communty Care Card Enr", $B1 = "Community Normal Enr"), OR($D1 = "Chisanga UHC", $D1 = "Kasama General", $D1 = "Kaizya HP", $D1 = "Kasakalawe HP", $D1 = "Mpulungu HAHC", $D1 = "Nsumbu RHC", $D1 = "Tulemane UHC"), OR($E1 = "MCH", $E1 = "Labour Ward", $E1 = "VCT", $E1 = "PITC", $E1 = "DCT", $E1 = "Fast Track", $E1 = "Traige", $E1 = "Youth Conner", $E1 = "OPD", $E1 = "VMMC", $E1 = "Indexing", $E1 = "ART", $E1 = "T.B", $E1 = "Community", $E1 = "Cervical Cancer", $E1 = "Mens Clinic", $E1 = "Pediatric Ward", $E1 = "PNS"), OR($M1 = "Male", $M1 = "Female"), OR($P1 = "Active", $P1 = "Inactive"), OR($Q1 = "Trans Out", $Q1 = "Trans In", $Q1 = "Deceased", $Q1 = "LTFU", $Q1 = "Deactivated", $Q1 = "Local"),  OR($AB1 = "YES", $AB1 = "NO", $AB1 = "NO Mobile #"), OR($AA1 = "Yes", $AA1 = "No"), OR($N1 = "English", $N1 = "Nyanja", $N1 = "Bemba", $N1 = "Lungu", $N1 = "Mambwe"), OR($S1 = "New", $S1 = "Old"), OR($BG1 = "Fishermen/women", $BG1 = "Farmers", $BG1 = "Traders", $BG1 = "Others", $BG1 = ""), OR($U1 = "Same in SC", $U1 = "Same in PRs", $U1 = "Same in Both", $U1 = "Different or No Address in PRs/Added", $U1 = "Different or No Address in Sc/Added", $U1 = "Different or No Address in Both/Added", $U1 = "Same in SC but Different or No Address in PRs", $U1 = "Same in PRs but Different or No Address in SC", $U1 = "Not Harmonized"), OR($AX1 = "Yes", $AX1 = "No"), OR($T1 = "Same in SC", $T1 = "Same in PRs", $T1 = "Same in Both", $T1 = "Different or No Mobile in PRs/Added", $T1 = "Different or No Mobile in Sc/Added", $T1 = "Different or No Mobile in Both/Added", $T1 = "Same in SC but Different or No Mobile in PRs", $T1 = "Same in PRs but Different or No Mobile in SC", $T1 = "Care Card", $T1 = "Not Harmonized"), OR($AW1 = "Not Eligible (TX_NEW)", $AW1 = "Results Found in SC and Updated in CTS", $AW1 = "Results Found in Physical Registers", $AW1 = "No VL Result found in SC or PRs", $AW1 = "VL Updated after follow up", $AW1 = "VL Results Pending Collection and Updates", $AW1 = ""), $A1 < TODAY(), $A1 > 01/01/2020, OR($C1 = "", AND($C1 <> "", $C1 > 1900-01-01), AND($C1 <> "", $C1 < TODAY())), $L1 > 01/01/1900, $L1 <= TODAY(), $AD1 < (TODAY() + 217), OR($AC1 = "", AND($AC1 <> "", $AC1 > (TODAY() - 217))), OR($AE1 = "", AND($AE1 <> "", $AE1 < (TODAY() + 217))), OR($AQ1 = "", AND($AQ1 = "", $BH1 <> "", $AW1 = "VL Results Pending Collection and Updates"), AND($AQ1 <> "", $AQ1 < (TODAY() + 730))), OR($BH1 = "", AND($BH1 <> "", $BH1 < TODAY())), OR($BD1 = "", AND($BD1 <> "", $BD1 < TODAY())))']
sheet.conditional_formatting.add(sheet.dimensions, rule)

#------------------------------------------------------------BORDERS ON ERRORED CELL--------------------------------------------------------------------------
thick_border = Side(border_style="thick")               #Defining thick borders
square_border = Border(top=thick_border, right=thick_border, bottom=thick_border, left=thick_border) #Defining border type as square


rows = list()                                           #Creating empty list to store data in each row
#newRows = list()                                        #New empty list to store data in each row for the report
dates = dict()                                          #Creating empty dictionay to store dates in each row
count = 1
newCount = 1                                            #New count for report data iteration                                        

reportDates = dict()
#------------------------------------------------------------Automating Creation of Dictionary---------------------------------------------

for row in sheet.iter_rows(min_row = 2,
                           min_col = 1,
                           values_only = True):         #To iterate through the workbook row-by-row
    rows = list()
    newCount += 1 

    for entry in row:
        rows.append(entry)

    dateCol = "A" + str(newCount    )                          #To produce cell incrementing cell numbers for column A
    dateVal = sheet[dateCol].value                      #Value of cells in Column A
    Enr_date = datetime.date(dateVal)                   #To convert datetime object to date

    reportDates [str(Enr_date)] = {}                    #Creating sub-dictionary to store data for each date    
    reportDates [str(Enr_date)]['pendingHarmonization PR'] = 0                        #To create a sub dictionary called pending harmonization and equate it to 0 for each date
    reportDates [str(Enr_date)]['pendingHarmonization SC'] = 0
    reportDates[str(Enr_date)]['Site Care Card'] = 0
    reportDates[str(Enr_date)]['Site Normal'] = 0
    reportDates[str(Enr_date)]['Communty Care Card'] = 0
    reportDates[str(Enr_date)]['Community Normal'] = 0
    reportDates[str(Enr_date)]['Reviewed # PR'] = 0
    reportDates[str(Enr_date)]['Reviewed # SC'] = 0
    reportDates[str(Enr_date)]['Reviewed add PR'] = 0
    reportDates[str(Enr_date)]['Reviewed add SC'] = 0

    reportDates[str(Enr_date)]['same # PR'] = 0
    reportDates[str(Enr_date)]['same add PR'] = 0
    reportDates[str(Enr_date)]['same # SC'] = 0
    reportDates[str(Enr_date)]['same add SC'] = 0
    reportDates[str(Enr_date)]['added # PR'] = 0
    reportDates[str(Enr_date)]['added add PR'] = 0
    reportDates[str(Enr_date)]['added # SC'] = 0
    reportDates[str(Enr_date)]['added add SC'] = 0
    
    
#--------------------------------------------------------------------------------------------------------------------------------------------


#Variable called count set at 1
for row in sheet.iter_rows(min_row = 2,
                           min_col = 1,
                           values_only = True):         #To iterate through the workbook row-by-row
    rows = list()
    count = count  + 1

    for entry in row:
        rows.append(entry)


#ORANGE FLAG

    #BORDERS ON COLUMN T|U (UNREVIEWED/UNHARMONIZED CELLS)

    column_T = 'T' + str(count)
    cell = sheet[column_T]
    
    if rows[19] == "Not Harmonized":
        cell.border = square_border                     #To apply boarders to cells in collumn T(row 19) that are "Not Harmonized"

    if rows[19] == None or rows[19] == "''" or rows[19] == "":
        cell.border = square_border                     #To apply boarders to cells in collumn T(row 19) that are Blank

    column_U = 'U' + str(count)
    cell = sheet[column_U]
    if rows[20] == "Not Harmonized":
        cell.border = square_border                     #To apply boarders to cells in collumn U(row 20) that are "Not Harmonized"

    if rows[20] == None or rows[20] == "''" or rows[20] == "":
        cell.border = square_border                     #To apply boarders to cells in collumn U(row 20) that are Blank
        

    #BORDERS ON COLUMN B AND COLUMN AA|AB (UNREVIEWED/UNHARMONIZED CELLS)
    #Site Normal
        
    column_AA = 'AA' + str(count)
    cell = sheet[column_AA]
    column_AB = 'AB' + str(count)
    cell_2 = sheet[column_AB]
    
    if rows[1] == "Site Normal Enr":
        if rows[26] == "No":
            cell.border = square_border
        if rows[27] == "NO":
            cell_2.border = square_border

    #Community Normal
            
    if rows[1] == "Community Normal Enr":
        if rows[26] == "No":
            cell.border = square_border
        if rows[27] == "NO":
            cell_2.border = square_border
    

     #BORDERS ON COLUMN AD(BLANK NEXT APP)
    column_AD = 'AD' + str(count)
    cell = sheet[column_AD]
    if rows[29] == None or rows[29] == "''" or rows[29] == "":
        cell.border = square_border                     #To apply boarders to cells in collumn AD(row 29) that are "Not Harmonized"    


#RED FLAG
        #---------------------------------------------------------DATES BORDERS (INCOMPLETE)------------------------------------------------------------------------
    cell_A = "A" + str(count)                            #To produce incrementing cell numbers for column A
    Val_A = sheet[cell_A].value                          #Value of cells in Column A
    Enr_date = datetime.date(Val_A)                      #To convert datetime object to date
    dates ["Enr_date"] = Enr_date                        #To add date object to dictionary

    cell_C = "C" + str(count)                            #To produce cell incrementing cell numbers for column K
    Val_C = sheet[cell_C].value                          #Value of cells in Column K
    if Val_C != None: 
        artStartDate = datetime.date(Val_C)              #To convert datetime object to date
    else:
        artStartDate = str(Val_C)
    dates ["ArtStartDate"] = artStartDate                #To add date object to dictionary
    
    cell_L = "L" + str(count)                            #To produce cell incrementing cell numbers for column K
    Val_L = sheet[cell_L].value                          #Value of cells in Column K
    DOB = datetime.date(Val_L)                           #To convert datetime object to date
    dates ["DOB"] = DOB                                  #To add date object to dictionary

    cell_R = "R" + str(count)                            #To produce cell incrementing cell numbers for column K
    Val_R = sheet[cell_R].value                          #Value of cells in Column K
    if Val_R != None: 
        statusID = datetime.date(Val_R)                  #To convert datetime object to date
    else:
        statusID = str(Val_R)
    dates ["Status Interaction Date"] = statusID         #To add date object to dictionary

    cell_AC = "AC" + str(count)                          #To produce cell incrementing cell numbers for column K
    Val_AC = sheet[cell_AC].value                        #Value of cells in Column K
    if Val_AC != None: 
        lastApt = datetime.date(Val_AC)                  #To convert datetime object to date
    else:
        lastApt = str(Val_AC)
    dates ["Last Apt"] = lastApt                         #To add date object to dictionary

    cell_AD = "AD" + str(count)                          #To produce cell incrementing cell numbers for column K
    Val_AD = sheet[cell_AD].value                        #Value of cells in Column K
    if Val_AD != None: 
        nxtApt = datetime.date(Val_AD)                  #To convert datetime object to date
    else:
        nxtApt = str(Val_AD)
    dates ["Next Apt"] = nxtApt                         #To add date object to dictionary

    cell_AE = "AE" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_AE = sheet[cell_AE].value                       #Value of cells in Column K
    if Val_AE != None: 
        revNxtApt = datetime.date(Val_AE)               #To convert datetime object to date
    else:
        revNxtApt = str(Val_AE)
    dates ["Revised Next Apt"] = revNxtApt              #To add date object to dictionary

    cell_AF = "AF" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_AF = sheet[cell_AF].value                       #Value of cells in Column K
    if Val_AF != None: 
        ap1 = datetime.date(Val_AF)                     #To convert datetime object to date
    else:
        ap1 = str(Val_AF)
    dates ["Ap1"] = ap1                                 #To add date object to dictionary


    cell_AG = "AG" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_AG = sheet[cell_AG].value                       #Value of cells in Column K
    if Val_AG != None: 
        ap2 = datetime.date(Val_AG)                     #To convert datetime object to date
    else:
        ap2 = str(Val_AG)
    dates ["Ap2"] = ap2                                 #To add date object to dictionary

    
    cell_AH = "AH" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_AH = sheet[cell_AH].value                       #Value of cells in Column K
    if Val_AH != None: 
        ap3 = datetime.date(Val_AH)                     #To convert datetime object to date
    else:
        ap3 = str(Val_AH)
    dates ["Ap3"] = ap3                                 #To add date object to dictionary

    
    cell_AI = "AI" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_AI = sheet[cell_AI].value                       #Value of cells in Column K
    if Val_AI != None: 
        ap4 = datetime.date(Val_AI)                     #To convert datetime object to date
    else:
        ap4 = str(Val_AI)
    dates ["Ap4"] = ap4                                 #To add date object to dictionary

    
    cell_AJ = "AJ" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_AJ = sheet[cell_AJ].value                       #Value of cells in Column K
    if Val_AJ != None: 
        ap5 = datetime.date(Val_AJ)                     #To convert datetime object to date
    else:
        ap5 = str(Val_AJ)
    dates ["Ap5"] = ap5                                 #To add date object to dictionary

    
    cell_AK = "AK" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_AK = sheet[cell_AK].value                       #Value of cells in Column K
    if Val_AK != None: 
        ap6 = datetime.date(Val_AK)                     #To convert datetime object to date
    else:
        ap6 = str(Val_AK)
    dates ["Ap6"] = ap6                                 #To add date object to dictionary

    
    cell_AL = "AL" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_AL = sheet[cell_AL].value                       #Value of cells in Column K
    if Val_AL != None: 
        ap7 = datetime.date(Val_AL)                     #To convert datetime object to date
    else:
        ap7 = str(Val_AL)
    dates ["Ap7"] = ap7                                 #To add date object to dictionary

    
    cell_AM = "AM" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_AM = sheet[cell_AM].value                       #Value of cells in Column K
    if Val_AM != None: 
        ap8 = datetime.date(Val_AM)                     #To convert datetime object to date
    else:
        ap8 = str(Val_AM)
    dates ["Ap8"] = ap8                                 #To add date object to dictionary

    
    cell_AN = "AN" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_AN = sheet[cell_AN].value                       #Value of cells in Column K
    if Val_AN != None: 
        ap9 = datetime.date(Val_AN)                     #To convert datetime object to date
    else:
        ap9 = str(Val_AN)
    dates ["Ap9"] = ap9                                 #To add date object to dictionary

    
    cell_AO = "AO" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_AO = sheet[cell_AO].value                       #Value of cells in Column K
    if Val_AO != None: 
        ap10 = datetime.date(Val_AO)                    #To convert datetime object to date
    else:
        ap10 = str(Val_AO)
    dates ["Ap10"] = ap10                               #To add date object to dictionary
    
    cell_AQ = "AQ" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_AQ = sheet[cell_AQ].value                       #Value of cells in Column K
    if Val_AQ != None: 
        vlDue = datetime.date(Val_AQ)                   #To convert datetime object to date
    else:
        vlDue = str(Val_AQ)
    dates ["VL Due"] = vlDue                            #To add date object to dictionary

    cell_BD = "BD" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_BD = sheet[cell_BD].value                       #Value of cells in Column K
    if Val_BD != None: 
        daySeen = datetime.date(Val_BD)                 #To convert datetime object to date
    else:
        daySeen = str(Val_BD)
    dates ["Actual Day Seen"] = daySeen                 #To add date object to dictionary

    cell_BD = "BD" + str(count)                         #To produce cell incrementing cell numbers for column K
    Val_BD = sheet[cell_BD].value                       #Value of cells in Column K
    if Val_BD != None: 
        vlDone = datetime.date(Val_BD)                  #To convert datetime object to date
    else:
        vlDone = str(Val_BD)
    dates ["VL Done"] = vlDone                          #To add date object to dictionary

    #-----------------------------------------------------------------------------------------------------------------------------------------------------------------

    #--------------------------------------------Boarders when Entry IS NOT one of the dropdown items provided--------------------------------------------------------------
    
    #BORDERS ON COLUMN B (Enr Type)
    column_B = 'B' + str(count) 
    cell_2 = sheet[column_B]

    if rows[1] != "Site Care Card Enr" and rows[1] != "Site Normal Enr"  and rows[1] != "Communty Care Card Enr" and rows[1] != "Community Normal Enr":
        cell_2.border = square_border


        #----------------------------------------COUNT FOR ENROLMENTS------------------------------

    if rows[1] == "Site Care Card Enr":
        reportDates[str(Enr_date)]['Site Care Card'] += 1
        
    if rows[1] == "Site Normal Enr":
        reportDates[str(Enr_date)]['Site Normal'] += 1

    if rows[1] == "Communty Care Card Enr":
        reportDates[str(Enr_date)]['Communty Care Card'] += 1

    if rows[1] == "Community Normal Enr":
        reportDates[str(Enr_date)]['Community Normal'] += 1
        #------------------------------------------------------------------------------------------------

    #BORDERS ON COLUMN D (Facility Name)
    column_D = 'D' + str(count) 
    cell_2 = sheet[column_D]    

    if rows[3] != "Chisanga UHC" and rows[3] != "Kasama General"  and rows[3] != "Kaizya HP" and rows[3] != "Kasakalawe HP" and rows[3] != "Mpulungu HAHC" and rows[3] != "Nsumbu RHC"  and rows[3] != "Tulemane UHC":
        cell_2.border = square_border

    #BORDERS ON COLUMN E (Department)
    column_E = 'E' + str(count) 
    cell_2 = sheet[column_E]    

    if rows[4] != "MCH" and rows[4] != "Labour Ward"  and rows[4] != "VCT" and rows[4] != "PITC" and rows[4] != "DCT" and rows[4] != "Fast Track" and rows[4] != "Traige" and rows[4] != "Youth Conner" and rows[4] != "OPD" and rows[4] != "VMMC" and rows[4] != "Indexing" and rows[4] != "ART" and rows[4] != "T.B" and rows[4] != "Community" and rows[4] != "Cervical Cancer" and rows[4] != "Mens Clinic" and rows[4] != "Chest Clinic" and rows[4] != "Pediatric Ward" and rows[4] != "PNS":
        cell_2.border = square_border


    #BORDERS ON COLUMN M (Sex)
    column_M = 'M' + str(count) 
    cell_2 = sheet[column_M]    

    if rows[12] != "Male" and rows[12] != "Female":
        cell_2.border = square_border


    #BORDERS ON COLUMN P (Status)
    column_P = 'P' + str(count) 
    cell_2 = sheet[column_P]    

    if rows[15] != "Active" and rows[15] != "Inactive":
        cell_2.border = square_border

    #BORDERS ON COLUMN Q (Status Comment)
    column_Q = 'Q' + str(count) 
    cell_2 = sheet[column_Q]    

    if rows[16] != "Trans Out" and rows[16] != "Trans In" and rows[16] != "Deceased" and rows[16] != "LTFU" and rows[16] != "Deactivated" and rows[16] != "Local":
        cell_2.border = square_border

    #BORDERS ON COLUMN AB (Verfied Mobile No)
    column_AB = 'AB' + str(count) 
    cell_2 = sheet[column_AB]    

    if rows[27] != "YES" and rows[27] != "NO" and rows[27] != "NO Mobile #":
        cell_2.border = square_border

    #BORDERS ON COLUMN AA (Internaal Mobile No)
    column_AA = 'AA' + str(count) 
    cell_2 = sheet[column_AA]    

    if rows[26] != "Yes" and rows[26] != "No":
        cell_2.border = square_border

    #BORDERS ON COLUMN N (Langueges)
    column_N = 'N' + str(count) 
    cell_2 = sheet[column_N]    

    if rows[13] != "English" and rows[13] != "Nyanja" and rows[13] != "Bemba" and rows[13] != "Lungu" and rows[13] != "Mambwe":
        cell_2.border = square_border

    #BORDERS ON COLUMN S (Client Type)
    column_S = 'S' + str(count) 
    cell_2 = sheet[column_S]    

    if rows[18] != "New" and rows[18] != "Old":
        cell_2.border = square_border

    #BORDERS ON COLUMN P (Client occupation)
    column_BG = 'BG' + str(count) 
    cell_2 = sheet[column_BG]    

    if rows[58] != "Fishermen/women" and rows[58] != "Farmers" and rows[58] != "Traders" and rows[58] != "Others" and rows[58] !=  None and rows[58] != "''" and rows[58] != "":
        cell_2.border = square_border

    #BORDERS ON COLUMN P (Harmonized residential address/Village/Township)
    column_P = 'P' + str(count) 
    cell_2 = sheet[column_P]    

    if rows[20] != "Same in SC" and rows[20] != "Same in PRs" and rows[20] != "Same in Both" and rows[20] != "Different or No Address in PRs/Added" and rows[20] != "Different or No Address in Sc/Added" and rows[20] != "Different or No Address in Both/Added" and rows[20] != "Same in SC but Different or No Address in PRs" and rows[20] != "Same in PRs but Different or No Address in SC" and rows[20] != "Not Harmonized":
        cell_2.border = square_border

    #BORDERS ON COLUMN AX (Address Impacted)
    column_AX = 'AX' + str(count) 
    cell_2 = sheet[column_AX]    

    if rows[49] != "Yes" and rows[49] != "No":
        cell_2.border = square_border

    #BORDERS ON COLUMN T (Harmonized Mobile #” )
    column_T = 'T' + str(count) 
    cell_2 = sheet[column_T]
    #print(cell_2.value)

    if rows[19] != "Same in SC" and rows[19] != "Same in PRs" and rows[19] != "Same in Both" and rows[19] != "Different or No Mobile in PRs/Added" and rows[19] != "Different or No Mobile in Sc/Added" and rows[19] != "Different or No Mobile in Both/Added" and rows[19] != "Same in SC but Different or No Mobile in PRs" and rows[19] != "Same in PRs but Different or No Mobile in SC" and rows[19] != "Care Card" and rows[19] != "Not Harmonized":
        cell_2.border = square_border


    #BORDERS ON COLUMN AW (VL Harmonization)
    column_AW = 'AW' + str(count) 
    cell_2 = sheet[column_AW]    

    if rows[48] != "Not Eligible (TX_NEW)" and rows[48] != "Results Found in SC and Updated in CTS" and rows[48] != "Results Found in Physical Registers" and rows[48] != "No VL Result found in SC or PRs" and rows[48] != "VL Updated after follow up" and rows[48] != "VL Results Pending Collection and Updates" and rows[48] !=  None and rows[48] != "''" and rows[48] != "":
        cell_2.border = square_border




    #------------------------------------------------------------------------REPORT COUNTS-----------------------------------------------------------

            #--------------------------------------------------------count for pending harmonization---------------------------------------------
        
    cell_A = "A" + str(count)                            #To produce cell incrementing cell numbers for column A
    Val_A = sheet[cell_A].value                          #Value of cells in Column A
    Enr_date = datetime.date(Val_A)

        
    if rows[19] == "Not Harmonized" or rows[20] == "Not Harmonized" or rows[19] == None or rows[20] == None or rows[19] == "''" or rows[20] == "''" or rows[19] == "" or rows[20] == "":
        reportDates[str(Enr_date)]['pendingHarmonization PR'] += 1  #TO add to pendingHarmonization PR count if entry is "Not Harmonized"
        reportDates[str(Enr_date)]['pendingHarmonization SC'] += 1  #TO add to pendingHarmonization SC count if entry is "Not Harmonized"
    
    if rows[19] == "Same in SC" or rows[20] == "Same in SC" or rows[19] == "Different or No Mobile in Sc/Added" or rows[20] == "Different or No Address in Sc/Added":
        reportDates[str(Enr_date)]['pendingHarmonization PR'] += 1

    if rows[19] == "Same in PRs" or rows[20] == "Same in PRs" or rows[19] == "Different or No Mobile in PRs/Added" or rows[20] == "Different or No Address in PRs/Added":
        reportDates[str(Enr_date)]['pendingHarmonization SC'] += 1
            
            #----------------------------------------------------------count for Reviewed-----------------------------------------------------------
        
            #Reviewed PR (Mobile & Address)
    if rows[19] != "Same in SC" and rows[19] != "Different or No Mobile in Sc/Added" and rows[19] != "Not Harmonized" and rows[19] != "Care Card":
        reportDates[str(Enr_date)]['Reviewed # PR'] += 1            #TO add to Reviewed # PR count if entry was reviewed in PR

    if rows[20] != "Same in SC" and rows[20] != "Different or No Address in Sc/Added" and rows[20] != "Not Harmonized":
        reportDates[str(Enr_date)]['Reviewed add PR'] += 1            #TO add to Reviewed # PR count if entry was reviewed in PR

            #Reviewed SC (Mobile & Address)
    if rows[19] != "Same in PRs" and rows[19] != "Different or No Mobile in PRs/Added" and rows[19] != "Not Harmonized" and rows[19] != "Care Card":
        reportDates[str(Enr_date)]['Reviewed # SC'] += 1            #TO add to Reviewed # PR count if entry was reviewed in PR

    if rows[20] != "Same in PRs" and rows[20] != "Different or No Address in PRs/Added" and rows[20] != "Not Harmonized":
        reportDates[str(Enr_date)]['Reviewed add SC'] += 1            #TO add to Reviewed # PR count if entry was reviewed in PR

            
            #------------------------------------------count for same & different mobiles & adresses in PR and SC-------------------------------------

            #same in SC (Mobile & Address)
    if rows[19] == "Same in SC" or rows[19] == "Same in Both" or rows[19] == "Same in SC but Different or No Mobile in PRs":
        reportDates[str(Enr_date)]['same # SC'] += 1            #TO add to same # PR count if entry was reviewed in PR

    if rows[20] == "Same in SC" or rows[20] == "Same in Both" or rows[20] == "Same in SC but Different or No Address in PRs":
        reportDates[str(Enr_date)]['same add SC'] += 1            #TO add to same add SC count if entry was reviewed in PR

            #same in PR (Mobile & Address)
    if rows[19] == "Same in PRs" or rows[19] == "Same in Both" or rows[19] == "Same in PRs but Different or No Mobile in SC":
        reportDates[str(Enr_date)]['same # PR'] += 1            #TO add to same # PR count if entry was reviewed in PR

    if rows[20] == "Same in PRs" or rows[20] == "Same in Both" or rows[20] == "Same in PRs but Different or No Address in SC":
        reportDates[str(Enr_date)]['same add PR'] += 1            #TO add to same add SC count if entry was reviewed in PR

            #added in SC (Mobile & Address)
    if rows[19] == "Different or No Mobile in Sc/Added" or rows[19] == "Different or No Mobile in Both/Added" or rows[19] == "Same in PRs but Different or No Mobile in SC":
        reportDates[str(Enr_date)]['added # SC'] += 1            #TO add to same # PR count if entry was reviewed in PR

    if rows[20] == "Different or No Address in Sc/Added" or rows[20] == "Different or No Address in Both/Added" or rows[20] == "Same in PRs but Different or No Address in SC":
        reportDates[str(Enr_date)]['added add SC'] += 1            #TO add to same add SC count if entry was reviewed in PR

            #added in PR (Mobile & Address)
    if rows[19] == "Different or No Mobile in PRs/Added" or rows[19] == "Different or No Mobile in Both/Added" or rows[19] == "Same in SC but Different or No Mobile in PRs":
        reportDates[str(Enr_date)]['added # PR'] += 1            #TO add to same # PR count if entry was reviewed in PR

    if rows[20] == "Different or No Address in PRs/Added" or rows[20] == "Different or No Address in Both/Added" or rows[20] == "Same in SC but Different or No Address in PRs":
        reportDates[str(Enr_date)]['added add PR'] += 1            #TO add to same add SC count if entry was reviewed in PR


    #-----------------------------------------------------------------IMPOSSIBLE DATES----------------------------------------------------------------------------

    #BORDERS ON COLUMN A (Enrollment Date)
    minEnrDate = date(year = 2020, month = 1, day = 1)     #Creating a 2020 date object
    column_A = 'A' + str(count) 
    cell_2 = sheet[column_A]

    if rows[0] > datetime.today() or datetime.date(rows[0]) < minEnrDate :
        cell_2.border = square_border


    #BORDERS ON COLUMN C (ART start date)
    minStartDate = date(year = 1900, month = 1, day = 1)     #Creating a 1900 date object
    column_C = 'C' + str(count) 
    cell_2 = sheet[column_C]

    if rows[2] !=  None and rows[2] != "''" and rows[2] != "":
        if rows[2] > datetime.today() or datetime.date(rows[2]) < minStartDate :
            cell_2.border = square_border

    #BORDERS ON COLUMN L (DOB)
    column_L = 'L' + str(count) 
    cell_2 = sheet[column_L]
    
    if rows[11] !=  None and rows[11] != "''" and rows[11] != "":
        if rows[11] > datetime.today() or datetime.date(rows[11]) < minStartDate :
            cell_2.border = square_border

    
    #BORDERS ON COLUMN AD (Next Apt)
    column_AD = 'AD' + str(count) 
    cell_2 = sheet[column_AD]
    delta = timedelta(days = 217)  #7 months is aproximately 217 days (7 months * 31 days)
    delta2 = today + delta         #Creatung a datetime object for 7 months from today
    if rows[29] !=  None and rows[29] != "''" and rows[29] != "":
        if datetime.date(rows[29]) > delta2:
            cell_2.border = square_border
            

    #BORDERS ON COLUMN AE (Revised Next Appointment)
    column_AE = 'AE' + str(count) 
    cell_2 = sheet[column_AE]
    if rows[30] !=  None and rows[30] != "''" and rows[30] != "":
        if datetime.date(rows[30]) > delta2:
            cell_2.border = square_border


    #BORDERS ON COLUMN AC (Last Appointment)
    column_AC = 'AC' + str(count) 
    cell_2 = sheet[column_AC]
    delta3 = today - delta         #Creatung a datetime object for 7 months before today
    if rows[28] !=  None and rows[28] != "''" and rows[28] != "":
        if datetime.date(rows[28]) < delta3:
            cell_2.border = square_border


    #BORDERS ON COLUMN AQ (VL Due Date)
    delta4 = timedelta(days = 730)      #2 years is aproximately 730 days (7 months * 31 days)
    delta5 = today + delta4             #Creatung a datetime object for 7 months from today
    column_AQ = 'AQ' + str(count) 
    cell_2 = sheet[column_AQ]
    if rows[42] !=  None and rows[42] != "''" and rows[42] != "":
        if datetime.date(rows[42]) > delta5:
            cell_2.border = square_border



    #BORDERS ON COLUMN BH (VL Done Date)
    column_BH = 'BH' + str(count) 
    cell_2 = sheet[column_BH]
    if rows[59] !=  None and rows[59] != "''" and rows[59] != "":
        if rows[59] > datetime.today():
            cell_2.border = square_border

    #BORDERS ON COLUMN BH & AQ (VL Done Date and VL Date Done)
    column_BH = 'BH' + str(count) 
    cell_2 = sheet[column_BH]
    cell_3 = sheet[column_AQ]
    if rows[59] !=  None and rows[59] != "''" and rows[59] != "" and rows[42] !=  None and rows[42] != "''" and rows[42] != "":
        if rows[59] == row[42]:
            cell_2.border = square_border
            cell_3.border = square_border

    if rows[59] !=  None and rows[59] != "''" and rows[59] != "" and rows[42] !=  None and rows[42] != "''" and rows[42] != "":
        if rows[59] > row[42]:
            cell_2.border = square_border
            cell_3.border = square_border
            


    #BORDERS ON COLUMN BD (Actual Day Seen at Facility)
    column_BD = 'BD' + str(count) 
    cell_2 = sheet[column_BD]
    if rows[55] !=  None and rows[55] != "''" and rows[55] != "":
        if rows[55] > datetime.today():
            cell_2.border = square_border


    #------------------------------------------------------------------BLANK CELLS/MISSING INFO----------------------------------------------------------------------
    cell_2 = sheet[column_A]
    if rows[0] ==  None or rows[0] == "''" or rows[0] == "":
        cell_2.border = square_border

    column_F = 'F' + str(count)
    cell_2 = sheet[column_F]
    if rows[5] ==  None or rows[5] == "''" or rows[5] == "":
        cell_2.border = square_border

    column_G = 'G' + str(count)
    cell_2 = sheet[column_G]
    if rows[6] ==  None or rows[6] == "''" or rows[6] == "":
        cell_2.border = square_border

    column_I = 'I' + str(count)
    cell_2 = sheet[column_I]
    if rows[8] ==  None or rows[8] == "''" or rows[8] == "":
        cell_2.border = square_border

    column_J = 'J' + str(count)
    cell_2 = sheet[column_J]
    if rows[9] ==  None or rows[9] == "''" or rows[9] == "":
        cell_2.border = square_border

    column_L = 'L' + str(count)
    cell_2 = sheet[column_L]
    if rows[11] ==  None or rows[11] == "''" or rows[11] == "":
        cell_2.border = square_border

    column_R = 'R' + str(count)
    cell_2 = sheet[column_R]
    if rows[17] ==  None or rows[17] == "''" or rows[17] == "":
        cell_2.border = square_border

    cell_3 = sheet[column_AW]
    if rows[19] != "Not Harmonized" and rows[20] != "Not Harmonized":
        if rows[48] ==  None or rows[48] == "''" or rows[48] == "":
            cell_3.border = square_border
        
#-------------------------------------------------------------------STATUS AND STATUS COMMENT--------------------------------------------------------------------

    column_P = 'P' + str(count)
    column_Q = 'Q' + str(count)
    cell_2 = sheet[column_P]
    cell_3 = sheet[column_Q]
    if rows[15] == "Active":
        if rows[16] != "Local" and rows[16] != "Trans In":
            cell_2.border = square_border
            cell_3.border = square_border

    if rows[15] == "Inactive":
        if rows[16] != "Trans Out" and rows[16] != "Deceased" and rows[16] != "LTFU" and rows[16] != "Deactivated":
            cell_2.border = square_border
            cell_3.border = square_border


#-------------------------------------------------------------------------CARE CARD---------------------------------------------------------------------------

    column_B = 'B' + str(count)
    column_AB = 'AB' + str(count)
    column_AA = 'AA' + str(count)
    cell_2 = sheet[column_B]
    cell_3 = sheet[column_AB]
    cell_4 = sheet[column_AA]
    
    if rows[1] == "Site Care Card Enr" or rows[1] == "Communty Care Card Enr":
        if rows[27] != "NO Mobile #":
            cell_2.border = square_border
            cell_3.border = square_border
        if rows[26] != "No":
            cell_2.border = square_border
            cell_4.border = square_border

    column_X = 'X' + str(count)
    column_Y = 'Y' + str(count)
    column_Z = 'Z' + str(count)
    cell_3 = sheet[column_X]
    cell_4 = sheet[column_Y]
    cell_5 = sheet[column_Z]
    if rows[1] == "Site Care Card Enr" or rows[1] == "Communty Care Card Enr":
        if rows[23] !=  None and rows[23] != "''" and rows[23] != "":
            cell_2.border = square_border
            cell_3.border = square_border
        if rows[24] !=  None and rows[24] != "''" and rows[24] != "":
            cell_2.border = square_border
            cell_4.border = square_border
        if rows[25] !=  None and rows[25] != "''" and rows[25] != "":
            cell_2.border = square_border
            cell_5.border = square_border

    column_T = 'T' + str(count)
    cell_3 = sheet[column_T]
    if rows[1] == "Site Care Card Enr" or rows[1] == "Communty Care Card Enr":
        if rows[19] != "Care Card" and rows[19] != "Not Harmonized":
            cell_2.border = square_border
            cell_3.border = square_border



#-------------------------------------------------------------------------NORMAL ENR---------------------------------------------------------------------------
    cell_3 = sheet[column_AB]
    if rows[1] == "Site Normal Enr" or rows[1] == "Community Normal Enr":
        if rows[27] == "NO Mobile #":
            cell_2.border = square_border
            cell_3.border = square_border

    cell_3 = sheet[column_X]
    cell_4 = sheet[column_Y]
    cell_5 = sheet[column_Z]
    if rows[1] == "Site Normal Enr" or rows[1] == "Community Normal Enr":
        if rows[19] != "Care Card":
            if rows[23] ==  None or rows[23] == "''" or rows[23] == "":
                if rows[24] ==  None or rows[24] == "''" or rows[24] == "":
                    if rows[25] ==  None or rows[25] == "''" or rows[25] == "":
                        cell_2.border = square_border
                        cell_3.border = square_border
                        cell_4.border = square_border
                        cell_5.border = square_border


#-----------------------------------------------------------------------------VL ENTRIES---------------------------------------------------------------------------
    column_AW = 'AW' + str(count)
    column_AR = 'AR' + str(count)
    column_AS = 'AS' + str(count)
    cell_3 = sheet[column_AW]
    cell_4 = sheet[column_AR]
    cell_5 = sheet[column_AS]
    cell_6 = sheet[column_BH]
    if rows[48] == "Results Found in Physical Registers" or rows[48] == "Results Found in SC and Updated in CTS":
        if rows[59] ==  None or rows[59] == "''" or rows[59] == "":
            cell_3.border = square_border
            cell_6.border = square_border
            
        if rows[43] ==  None or rows[43] == "''" or rows[43] == "":
                if rows[44] ==  None or rows[44] == "''" or rows[44] == "":
                    cell_3.border = square_border
                    cell_4.border = square_border
                    cell_5.border = square_border
      

#-----------------------------------------------------------------------------MOBILE NUMBERS---------------------------------------------------------------------------
    cell_3 = sheet[column_X]
    cell_4 = sheet[column_Y]
    cell_5 = sheet[column_Z]
    if rows[23] !=  None and rows[23] != "''" and rows[23] != "":
        if str(rows[23])[:2] != '77' and str(rows[23])[:2] != '97':
            cell_3.border = square_border

    if rows[24] !=  None and rows[24] != "''" and rows[24] != "":
        if str(rows[24])[:2] != '75' and str(rows[24])[:2] != '95':
            cell_4.border = square_border

    if rows[25] !=  None and rows[25] != "''" and rows[25] != "":
        if str(rows[25])[:2] != '76' and str(rows[25])[:2] != '96':            
            cell_5.border = square_border

    
        
       
#-----------------------------------------------------------------------------ADDRESS IMPACTED---------------------------------------------------------------------------
    column_AX = 'AX' + str(count)
    column_U = 'U' + str(count)
    cell_3 = sheet[column_AX]
    cell_4 = sheet[column_U]

    if rows[19] == "Not Harmonized" or rows[19] ==  None or rows[19] == "''" or rows[19] == "":
        if rows[49] == "Yes":
            cell_3.border = square_border
            cell_4.border = square_border
    


#----------------------------------------------------------------------------------------------------------------------------------------------------------------

sheet.freeze_panes = "A2" #To freeze first row
sheet.auto_filter.ref = sheet.dimensions    #To add filters
#sheet.auto_filter.add_sort_condition("$A2") To sort data by enr date
sheet.column_dimensions["T"].width = 15.43   #Slightly increase width of column for mobile harmonization
sheet.column_dimensions["U"].width = 15.43   ##Slightly increase width of column for residential harmonization

#-----------------------------------------------------------------------------REPORT-----------------------------------------------------------------------------

workbook.create_sheet("Report", 2)

sheet2 = workbook.worksheets[1]
sheet2["A1"] = "DATE"
sheet2.merge_cells('A1:A2') #Merge date columns
sheet2["B1"] = "SITE"
sheet2.merge_cells('B1:D1') #Merge site columns
sheet2["E1"] = "COMMUNITY"  
sheet2.merge_cells('E1:G1') #Merge community columns
sheet2["H1"] = "TOTAL" 
sheet2.merge_cells('H1:J1') #Merge total columns
sheet2["K1"] = "MOBILE # HARMONIZATION"
sheet2.merge_cells('K1:P1') #Merge mobile # harmonization columns
sheet2["Q1"] = "RESIDENTIAL ADDRESS HARMONIZATION"
sheet2.merge_cells('Q1:V1') #Merge address harmonization columns
sheet2["W1"] = "PENDING HARMONIZATION"
sheet2.merge_cells('W1:X1') #Merge total harmonized cells
sheet2["Y1"] = "CUMMULATIVE PENDING HARMONIZATION" 
sheet2.merge_cells('Y1:Z1') #Merge pending harmonization cells


sheet2["A1"].font = Font(bold=True)
sheet2["B1"].font = Font(bold=True)
sheet2["E1"].font = Font(bold=True)
sheet2["H1"].font = Font(bold=True)
sheet2["K1"].font = Font(bold=True)
sheet2["T1"].font = Font(bold=True)
sheet2["Q1"].font = Font(bold=True)
sheet2["W1"].font = Font(bold=True)
sheet2["Y1"].font = Font(bold=True)
sheet2["AA1"].font = Font(bold=True)
sheet2["A1"].alignment = Alignment(vertical='center', wrapText=True, horizontal='center')
sheet2["B1"].alignment = Alignment(vertical='center', wrapText=True, horizontal='center')
sheet2["E1"].alignment = Alignment(vertical='center', wrapText=True, horizontal='center')
sheet2["H1"].alignment = Alignment(vertical='center', wrapText=True, horizontal='center')
sheet2["K1"].alignment = Alignment(vertical='center', wrapText=True, horizontal='center')
sheet2["T1"].alignment = Alignment(vertical='center', wrapText=True, horizontal='center')
sheet2["Q1"].alignment = Alignment(vertical='center', wrapText=True, horizontal='center')
sheet2["W1"].alignment = Alignment(vertical='center', wrapText=True, horizontal='center')
sheet2["Y1"].alignment = Alignment(vertical='center', wrapText=True, horizontal='center')


sheet2["B2"] = "SITE NORMAL ENROLEMENTS"
sheet2["C2"] = "SITE CARECRD ENROLEMENTS"
sheet2["D2"] = "TOTAL SITE ENROLEMENTS"
sheet2["E2"] = "COMMUNITY NORMAL ENROLEMENTS"
sheet2["F2"] = "COMMUNITY CARECARD ENROLEMENTS"
sheet2["G2"] = "TOTAL COMMUNITY ENROLMENTS"
sheet2["H2"] = "NORMAL ENROLEMENTS"
sheet2["I2"] = "CARECARD ENROLEMENTS"
sheet2["J2"] = "TOTAL ENROLEMENTS"
sheet2["K2"] = "REVIEWED # IN PR"
sheet2["L2"] = "SAME # IN PR"
sheet2["M2"] = "ADDED # IN PR"
sheet2["N2"] = "REVIEWED # IN SC"
sheet2["O2"] = "SAME # IN SC"
sheet2["P2"] = "ADDED # IN SC"
sheet2["Q2"] = "REVIEWED ADDRESS IN PR"
sheet2["R2"] = "SAME ADDRESS IN PR"
sheet2["S2"] = "ADDED ADDRESS IN PR"
sheet2["T2"] = "REVIEWED ADDRESS in SC"
sheet2["U2"] = "SAME ADDRESS IN SC"
sheet2["V2"] = "ADDED ADDRESS IN SC"
sheet2["W2"] = "PR"
sheet2["X2"] = "SC"
sheet2["Y2"] = "PR"
sheet2["Z2"] = "SC"


sheet2.freeze_panes = "B3" #To freeze first row
#sheet2.auto_filter.ref = sheet.dimensions #To add filters


for col in sheet2.iter_cols(min_row = 2,
                           min_col = 2):
    cell = str(col).split('.')[1].split('>')[0]
    sheet2[cell].font = Font(bold=True)
    sheet2[cell].alignment = Alignment(vertical='center', wrapText=True, horizontal='center')
    sheet2.column_dimensions[cell[:-1]].width = 15
    
sheet2.column_dimensions["AE"].width = 17
sheet2.column_dimensions["AA"].width = 17
#Report Dates
sheet2.column_dimensions["A"].width = 15

newCount = 3
col_A = 'A'         #Date column
col_B = 'B'
col_C = 'C'
col_D = 'D'         #Total site enrollments
col_E = 'E'
col_F = 'F'
col_G = 'G'         #Total Community enrollments
col_H = 'H'         #Total Normal Enr (Site + Community)
col_I = 'I'         #Total Carecard Enr (Site + Community)
col_J = 'J'         #Total Enr (Site + Community)
col_K = 'K'         
col_L = 'L'
col_M = 'M'
col_N = 'N'
col_O = 'O'
col_P = 'P'
col_Q = 'Q'
col_R = 'R'
col_S = 'S'
col_T = 'T'
col_U = 'U'
col_V = 'V'
col_W = 'W'         #Pending Harmonization PR column
col_X = 'X'         #Pending Harmonization SC column
col_Y = 'Y'         #Cummulative pending Harmonization PR column
col_Z = 'Z'         #Cummulative pending Harmonization SC column

#------------------------------------------Cummulative pending harmonization----------------------------------

pendingPR = []                                                #Creating empty list to store pending harmonization numbers
pendingSC = []

for i in list(reportDates):
    pendingPR.append(reportDates[i]['pendingHarmonization PR'])  #Adding the pending harmonization numbers to pending list
    pendingSC.append(reportDates[i]['pendingHarmonization SC'])  #Adding the pending harmonization numbers to pending list

pendingPR.reverse()                                           #Reversing list so that they may be in accending order
pendingSC.reverse()                                           #Reversing list so that they may be in accending order

cummSumPR = []                                                #Creating list to store cummulative list
cummSumSC = []

x = 0                                                       #Creating variable to store cummulative values
for i in pendingPR:
    x+=i
    cummSumPR.append(x)

cummSumPR.reverse()                                           #Reaversing list back to decending order

y = 0                                                       #Creating variable to store cummulative values

for i in pendingSC:
    y+=i
    cummSumSC.append(y)

cummSumSC.reverse()                                           #Reaversing list back to decending order

cummCount = 0                                               #Creating a count to help with iteration

#-------------------------------------------------------------------------------------------------------------    

#------------------------------------------------------------------REPORT AESTHETICS------------------------------------------------------------------------    
thin_border = Side(border_style="thin") 
left_border = Border(left=thin_border)
right_border = Border(right=thin_border)
bottom_right_border = Border(bottom=thin_border, right=thin_border)
top_border = Border(top=thin_border)
sides_border = Border(right=thin_border, left=thin_border)


sheet2['D2'].fill = PatternFill("solid", start_color="D3D3D3")                  #Light-Grey fill
sheet2['D2'].border = sides_border
sheet2['G2'].fill = PatternFill("solid", start_color="D3D3D3")                  #Light-Grey fill
sheet2['G2'].border = sides_border
sheet2['J2'].fill = PatternFill("solid", start_color="D3D3D3")                  #Light-Grey fill
sheet2['J2'].border = sides_border
sheet2['K2'].fill = PatternFill("solid", start_color="CCDFE5")                  #Blue-Grey fill
sheet2['K2'].border = sides_border
sheet2['N2'].fill = PatternFill("solid", start_color="CCDFE5")                  #Blue-Grey fill
sheet2['N2'].border = sides_border
sheet2['Q2'].fill = PatternFill("solid", start_color="CCDFE5")                  #Blue-Grey fill
sheet2['Q2'].border = sides_border
sheet2['T2'].fill = PatternFill("solid", start_color="CCDFE5")                  #Blue-Grey fill
sheet2['T2'].border = sides_border
sheet2['W2'].fill = PatternFill("solid", start_color="FC957F")                  #Light Red Fill
sheet2['W2'].border = left_border
sheet2['X2'].fill = PatternFill("solid", start_color="FC957F")                  #Light Red Fill
sheet2['Y2'].fill = PatternFill("solid", start_color="FC957F")                  #Light Red Fill
sheet2['Y2'].border = left_border
sheet2['Z2'].fill = PatternFill("solid", start_color="FC957F")                  #Light Red Fill
sheet2['Z2'].border = right_border
for col in sheet2.iter_cols(min_row = 1,
                            max_row = 1,
                            min_col = 2,
                            max_col = 26):
    cell = str(col).split('.')[1].split('>')[0]
    sheet2[cell].border = right_border

    

for col in sheet2.iter_cols(min_row = 1,
                            max_row = 1,
                            min_col = 2,
                            max_col = 26):
    cell = str(col).split('.')[1].split('>')[0]
    sheet2[cell].border = bottom_right_border
       
#------------------------------------------------------------------------------------------------------------------------------------------------------------    


for i in list(reportDates):
    #Write Date
    cell = col_A + str(newCount)
    a = i.split('-')
    a.reverse()
    a = '-'.join(a)
    sheet2[cell] = a
    sheet2[cell].alignment = Alignment(horizontal='center')


    #Write Site Normal Enrollments
    cell = col_B + str(newCount)
    sheet2[cell] = reportDates[i]['Site Normal']
    sheet2[cell].alignment = Alignment(horizontal='center')
    
    
    #Write Site CareCard Enrollments
    cell = col_C + str(newCount)
    sheet2[cell] = reportDates[i]['Site Care Card']
    sheet2[cell].alignment = Alignment(horizontal='center')

    #Write Total Site Enrollments
    cell = col_D + str(newCount)
    sheet2[cell] = reportDates[i]['Site Normal'] + reportDates[i]['Site Care Card']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].fill = PatternFill("solid", start_color="D3D3D3")                  #Light-Grey fill
    sheet2[cell].border = left_border

    #Write Community Normal Enrollments
    cell = col_E + str(newCount)
    sheet2[cell] = reportDates[i]['Community Normal']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].border = left_border
    
    #Write Community CareCard Enrollments
    cell = col_F + str(newCount)
    sheet2[cell] = reportDates[i]['Communty Care Card']
    sheet2[cell].alignment = Alignment(horizontal='center')

    #Write Total Community Enrollments
    cell = col_G + str(newCount)
    sheet2[cell] = reportDates[i]['Community Normal'] + reportDates[i]['Communty Care Card'] 
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].fill = PatternFill("solid", start_color="D3D3D3")                  #Light-Grey fill
    sheet2[cell].border = left_border
    
    
    #Write Total Normal Enrollments
    cell = col_H + str(newCount)
    sheet2[cell] = reportDates[i]['Site Normal'] + reportDates[i]['Community Normal']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].border = left_border
    

    #Write Total CareCard Enrollments
    cell = col_I + str(newCount)
    sheet2[cell] = reportDates[i]['Site Care Card'] + reportDates[i]['Communty Care Card']
    sheet2[cell].alignment = Alignment(horizontal='center')

    #Write Total Enrollments
    cell = col_J + str(newCount)
    sheet2[cell] = reportDates[i]['Site Normal'] + reportDates[i]['Community Normal'] + reportDates[i]['Site Care Card'] + reportDates[i]['Communty Care Card']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].fill = PatternFill("solid", start_color="D3D3D3")                  #Light-Grey fill
    sheet2[cell].border = left_border

    #Write Reviewed # in PR 
    cell = col_K + str(newCount)
    sheet2[cell] = reportDates[i]['Reviewed # PR']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].fill = PatternFill("solid", start_color="CCDFE5")                  #Blue-Grey fill
    sheet2[cell].border = left_border

    #Write same # in PR 
    cell = col_L + str(newCount)
    sheet2[cell] = reportDates[i]['same # PR']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].border = left_border

    #Write added # in PR 
    cell = col_M + str(newCount)
    sheet2[cell] = reportDates[i]['added # PR']
    sheet2[cell].alignment = Alignment(horizontal='center')

    #Write Reviewed # in SC 
    cell = col_N + str(newCount)
    sheet2[cell] = reportDates[i]['Reviewed # SC']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].fill = PatternFill("solid", start_color="CCDFE5")                  #Blue-Grey fill
    sheet2[cell].border = left_border

    #Write same # in SC 
    cell = col_O + str(newCount)
    sheet2[cell] = reportDates[i]['same # SC']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].border = left_border

    #Write added # in SC 
    cell = col_P + str(newCount)
    sheet2[cell] = reportDates[i]['added # SC']
    sheet2[cell].alignment = Alignment(horizontal='center')

    #Write Reviewed add in PR 
    cell = col_Q + str(newCount)
    sheet2[cell] = reportDates[i]['Reviewed add PR']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].fill = PatternFill("solid", start_color="CCDFE5")                  #Blue-Grey fill
    sheet2[cell].border = left_border

    #Write same add in PR 
    cell = col_R + str(newCount)
    sheet2[cell] = reportDates[i]['same add PR']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].border = left_border

    #Write added add in PR 
    cell = col_S + str(newCount)
    sheet2[cell] = reportDates[i]['added add PR']
    sheet2[cell].alignment = Alignment(horizontal='center')

    #Write Reviewed add in SC 
    cell = col_T + str(newCount)
    sheet2[cell] = reportDates[i]['Reviewed add SC']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].fill = PatternFill("solid", start_color="CCDFE5")                  #Blue-Grey fill
    sheet2[cell].border = left_border

    #Write same add in SC 
    cell = col_U + str(newCount)
    sheet2[cell] = reportDates[i]['same add SC']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].border = left_border

    #Write added # in SC 
    cell = col_V + str(newCount)
    sheet2[cell] = reportDates[i]['added add SC']
    sheet2[cell].alignment = Alignment(horizontal='center')

    #Write Pending PR
    cell = col_W + str(newCount)
    sheet2[cell] = reportDates[i]['pendingHarmonization PR']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].fill = PatternFill("solid", start_color="FC957F")                  #Light Red Fill
    sheet2[cell].border = left_border
    
    #Write Pending SC
    cell = col_X + str(newCount)
    sheet2[cell] = reportDates[i]['pendingHarmonization SC']
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].fill = PatternFill("solid", start_color="FC957F")                  #Light Red Fill

    #Write Cummulative Pending PR
    cell = col_Y + str(newCount)
    sheet2[cell] = cummSumPR[cummCount]
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].fill = PatternFill("solid", start_color="FC957F")                  #Light Red Fill
    sheet2[cell].border = left_border

    #Write Cummulative Pending SC
    cell = col_Z + str(newCount)
    sheet2[cell] = cummSumSC[cummCount]
    sheet2[cell].alignment = Alignment(horizontal='center')
    sheet2[cell].fill = PatternFill("solid", start_color="FC957F")                  #Light Red Fill
    sheet2[cell].border = right_border

    
    newCount += 1
    cummCount += 1  
    

#----------------------------------------------------------------------------------------------------------------------------------------------------------------

workbook.save(new_file_name)

window = tk.Tk()
window.withdraw()
messagebox.showinfo('Success', "DQA complete")
new_file_path = os.path.join(cwd, new_file_name)
cmd = new_file_path
fp = os.popen(cmd)   #To launch the above command. The argument is a string that contains a shell command.
print(fp.read())     #To read the output of the object
finished = fp.close()   #command to close fp pipeline
